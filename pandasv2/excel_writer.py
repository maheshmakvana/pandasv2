"""
ExcelWriter support for pandasv2.

Provides a full-featured ExcelWriter context manager that wraps
pandas.ExcelWriter with additional helpers:
  - write_dataframe()    — write a DataFrame to a named sheet
  - write_many()         — write multiple DataFrames in one call
  - write_with_style()   — write with optional conditional formatting
  - sheet_names          — list sheets written so far
  - ExcelWriter          — drop-in replacement context manager

Built by Mahesh Makvana
"""

import pandas as _pd
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union


class ExcelWriter:
    """
    pandasv2 ExcelWriter — context manager for writing DataFrames to Excel.

    Drop-in replacement for pandas.ExcelWriter with extra helpers.

    Usage:
        >>> import pandasv2 as pd
        >>>
        >>> # Basic — identical to pandas
        >>> with pd.ExcelWriter('output.xlsx') as writer:
        ...     df1.to_excel(writer, sheet_name='Sales')
        ...     df2.to_excel(writer, sheet_name='Inventory')
        >>>
        >>> # Extra helpers
        >>> with pd.ExcelWriter('report.xlsx') as writer:
        ...     writer.write_dataframe(df, 'Sheet1', index=False)
        ...     writer.write_many({'Sales': df1, 'Stock': df2})
    """

    def __init__(
        self,
        path,
        engine=None,
        date_format=None,
        datetime_format=None,
        mode='w',
        storage_options=None,
        if_sheet_exists=None,
        engine_kwargs=None,
        **kwargs,
    ):
        """
        Initialize ExcelWriter.

        Args:
            path: File path or writable object for the Excel file
            engine: Engine to use ('openpyxl', 'xlsxwriter', 'odf')
            date_format: Format string for dates (default: 'YYYY-MM-DD')
            datetime_format: Format string for datetimes
            mode: 'w' (write, default) or 'a' (append)
            storage_options: Dict of storage options for cloud paths
            if_sheet_exists: 'error'|'new'|'replace'|'overlay' (append mode)
            engine_kwargs: Keyword arguments passed to the engine constructor
        """
        kw = dict(mode=mode)
        if engine is not None:
            kw['engine'] = engine
        if date_format is not None:
            kw['date_format'] = date_format
        if datetime_format is not None:
            kw['datetime_format'] = datetime_format
        if storage_options is not None:
            kw['storage_options'] = storage_options
        if if_sheet_exists is not None:
            kw['if_sheet_exists'] = if_sheet_exists
        if engine_kwargs is not None:
            kw['engine_kwargs'] = engine_kwargs
        kw.update(kwargs)

        self._writer = _pd.ExcelWriter(path, **kw)
        self._sheets_written: List[str] = []

    # ------------------------------------------------------------------
    # Context manager protocol
    # ------------------------------------------------------------------

    def __enter__(self) -> 'ExcelWriter':
        self._writer.__enter__()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        return self._writer.__exit__(exc_type, exc_val, exc_tb)

    # ------------------------------------------------------------------
    # Standard pandas.ExcelWriter interface
    # ------------------------------------------------------------------

    def __getattr__(self, name):
        """Delegate any unknown attribute to the underlying pandas writer."""
        return getattr(self._writer, name)

    @property
    def book(self):
        """Underlying workbook object (engine-specific)."""
        return self._writer.book

    @property
    def sheets(self):
        """Dict of sheet_name → worksheet objects."""
        return self._writer.sheets

    @property
    def path(self):
        """File path being written to."""
        return self._writer.path

    def save(self):
        """Save and close the workbook."""
        return self._writer.close()

    def close(self):
        """Close and save the workbook."""
        return self._writer.close()

    # ------------------------------------------------------------------
    # pandasv2 extras
    # ------------------------------------------------------------------

    @property
    def sheet_names(self) -> List[str]:
        """List of sheet names written so far in this session."""
        return list(self._sheets_written)

    def write_dataframe(
        self,
        df: _pd.DataFrame,
        sheet_name: str = 'Sheet1',
        startrow: int = 0,
        startcol: int = 0,
        index: bool = True,
        header: bool = True,
        float_format: Optional[str] = None,
        columns: Optional[List[str]] = None,
        na_rep: str = '',
        inf_rep: str = '',
        freeze_panes: Optional[Tuple[int, int]] = None,
        **kwargs,
    ) -> None:
        """
        Write a DataFrame to a named sheet with common options.

        Args:
            df: DataFrame to write
            sheet_name: Target sheet name
            startrow: Row offset (0-indexed)
            startcol: Column offset (0-indexed)
            index: Write row index
            header: Write column headers
            float_format: Format string for floats (e.g. '%.2f')
            columns: Subset of columns to write
            na_rep: String representation for NaN
            inf_rep: String representation for Inf
            freeze_panes: (row, col) to freeze panes at
            **kwargs: Extra kwargs passed to DataFrame.to_excel

        Example:
            >>> with pd.ExcelWriter('out.xlsx') as writer:
            ...     writer.write_dataframe(df, 'Sales', index=False,
            ...                            float_format='%.2f',
            ...                            freeze_panes=(1, 0))
        """
        kw = dict(
            excel_writer=self._writer,
            sheet_name=sheet_name,
            startrow=startrow,
            startcol=startcol,
            index=index,
            header=header,
            na_rep=na_rep,
            inf_rep=inf_rep,
        )
        if float_format is not None:
            kw['float_format'] = float_format
        if columns is not None:
            kw['columns'] = columns
        if freeze_panes is not None:
            kw['freeze_panes'] = freeze_panes
        kw.update(kwargs)

        df.to_excel(**kw)
        if sheet_name not in self._sheets_written:
            self._sheets_written.append(sheet_name)

    def write_many(
        self,
        frames: Dict[str, _pd.DataFrame],
        index: bool = False,
        **kwargs,
    ) -> None:
        """
        Write multiple DataFrames — one per sheet — in a single call.

        Args:
            frames: Dict of {sheet_name: DataFrame}
            index: Whether to include the index (default False)
            **kwargs: Extra kwargs passed to each write_dataframe call

        Example:
            >>> with pd.ExcelWriter('report.xlsx') as writer:
            ...     writer.write_many({
            ...         'Sales':     sales_df,
            ...         'Inventory': stock_df,
            ...         'Returns':   returns_df,
            ...     })
        """
        for sheet_name, df in frames.items():
            self.write_dataframe(df, sheet_name=sheet_name, index=index, **kwargs)

    def write_with_style(
        self,
        df: _pd.DataFrame,
        sheet_name: str = 'Sheet1',
        header_fill: Optional[str] = None,
        header_font_bold: bool = True,
        col_widths: Optional[Dict[str, int]] = None,
        index: bool = False,
        **kwargs,
    ) -> None:
        """
        Write DataFrame with basic styling (requires openpyxl).

        Args:
            df: DataFrame to write
            sheet_name: Target sheet name
            header_fill: Hex colour for header cells (e.g. '4472C4')
            header_font_bold: Bold header text
            col_widths: Dict of {column_name: width} overrides
            index: Write row index
            **kwargs: Extra kwargs passed to write_dataframe

        Example:
            >>> with pd.ExcelWriter('styled.xlsx', engine='openpyxl') as writer:
            ...     writer.write_with_style(
            ...         df, 'Report',
            ...         header_fill='4472C4',
            ...         col_widths={'Name': 25, 'Value': 15},
            ...     )
        """
        self.write_dataframe(df, sheet_name=sheet_name, index=index, **kwargs)

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            return  # silently skip styling if openpyxl not installed

        ws = self._writer.sheets[sheet_name]
        n_index_cols = df.index.nlevels if index else 0
        n_cols = len(df.columns) + n_index_cols

        # Style header row
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            if header_fill:
                cell.fill = PatternFill(
                    start_color=header_fill.lstrip('#'),
                    end_color=header_fill.lstrip('#'),
                    fill_type='solid',
                )
            if header_font_bold:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Auto-fit column widths
        for i, col in enumerate(df.columns):
            col_letter = openpyxl.utils.get_column_letter(i + 1 + n_index_cols)
            user_width = (col_widths or {}).get(col)
            if user_width:
                ws.column_dimensions[col_letter].width = user_width
            else:
                max_len = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max() if len(df) > 0 else 0,
                )
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def auto_fit_columns(self, sheet_name: str) -> None:
        """
        Auto-fit column widths for a sheet (requires openpyxl).

        Must be called after write_dataframe/write_many.

        Example:
            >>> with pd.ExcelWriter('out.xlsx', engine='openpyxl') as writer:
            ...     writer.write_dataframe(df, 'Data')
            ...     writer.auto_fit_columns('Data')
        """
        try:
            import openpyxl
        except ImportError:
            return

        ws = self._writer.sheets.get(sheet_name)
        if ws is None:
            return

        for col_cells in ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
